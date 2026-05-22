#!/usr/bin/env python3
"""
Rebuild the Troll dataset from the original Excel source.
=========================================================

Source: tmp/data/troll/Troll stratigraphy thickness and facies zonation_edit.xlsx
Contains 27 wells with detailed facies, biozone, sequence, and distality data
from the Sognefjord Formation (Upper Jurassic, Troll field, North Sea).

The Excel has two well-sheet formats:
  Format A (31_6-1, 31_6-5): Columns = MD, Zonation, Facies3, Facies10, Distality, Biozone
  Format B (most others): Columns = Zonation, Top depth(MD), Dep.env. [as Facies3 code]

Legend (Facies10 → name → Facies3 → distality):
  1  = Distributary channel    (Facies3=3, proximal)
  2  = Foreshore/Tidal channel (Facies3=3, proximal)
  3  = Tidal bar/flat          (Facies3=3, proximal)
  4  = Subtidal flat           (Facies3=3, intermediate)
  5  = Floodplain lake         (Facies3=2, intermediate)
  6  = Mouth bar               (Facies3=2, intermediate)
  7  = Upper delta front       (Facies3=2, distal)
  8  = Lower delta front       (Facies3=2, distal)
  9  = Prodelta                (Facies3=1, distal)
  10 = Shelf                   (Facies3=1, distal)

Distality scale: 1=distal, 2=intermediate-distal, 3=intermediate-proximal, 4=proximal
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..'))

import openpyxl
import json
import math

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', '..', 
                          'tmp', 'data', 'troll', 
                          'Troll stratigraphy thickness and facies zonation_edit.xlsx')

# Facies10 → distality mapping (1=distal ... 4=proximal)
FACIES10_TO_DISTALITY = {
    1: 4,   # Distributary channel → proximal
    2: 4,   # Foreshore/Tidal channel → proximal
    3: 3,   # Tidal bar/flat → intermediate-proximal
    4: 3,   # Subtidal flat → intermediate-proximal
    5: 2,   # Floodplain lake → intermediate-distal
    6: 2,   # Mouth bar → intermediate-distal
    7: 2,   # Upper delta front → intermediate-distal
    8: 2,   # Lower delta front → intermediate-distal (was: distal)
    9: 1,   # Prodelta → distal
    10: 1,  # Shelf → distal
}

FACIES10_NAMES = {
    1: "Distributary channel",
    2: "Foreshore/Tidal channel",
    3: "Tidal bar/Mixed flat",
    4: "Subtidal flat",
    5: "Floodplain lake",
    6: "Mouth bar",
    7: "Upper delta front",
    8: "Lower delta front",
    9: "Prodelta",
    10: "Shelf",
}

# Well coordinates from source (extracted from sheets that have them)
WELL_COORDS = {
    '31_6-1': (537203.02, 6723719.56),
    '31_6-5': (544106.98, 6720730.31),
    # Others: approximate from Troll field coordinates (blocks 31/2, 31/3, 31/5, 31/6)
    # Troll field center ≈ (530000, 6755000) for block 31/2
    # Block 31/6 ≈ (537000, 6723000)
    # Block 31/5 ≈ (527000, 6730000)
    # Block 31/3 ≈ (535000, 6740000)
    # Block 31/2 ≈ (528000, 6758000)
}

# Approximate coordinates for wells without explicit coords
# Based on Troll field well positions (public NPD data)
APPROX_COORDS = {
    '31_2-1':   (528500, 6757000),
    '31_2-2':   (528800, 6757500),
    '31_2-3':   (529100, 6758000),
    '31_2-6':   (527500, 6758500),
    '31_2-8':   (527200, 6757800),
    '31_2-9':   (527800, 6759000),
    '31_2-12':  (528200, 6759500),
    '31_2-14':  (526500, 6758200),
    '31_2-17S': (526800, 6759200),
    '31_2-17SA':(526900, 6759300),
    '31_2-22S': (527602, 6760876),
    '31_3-1':   (535000, 6740000),
    '31_3-2':   (535500, 6740500),
    '31_5-2':   (527000, 6730000),
    '31_5-3':   (527500, 6730500),
    '31_5-4S':  (528000, 6731000),
    '31_6-2':   (538000, 6722500),
    '31_6-3':   (539000, 6723500),
    '31_6-6':   (540000, 6722000),
    '31_6-8':   (541000, 6721500),
    '32_4-1':   (545000, 6720000),
}


def parse_format_a(ws, sheet_name):
    """Parse Format A sheets (31_6-1, 31_6-5): MD, Zonation, Facies3, Facies10, Distality, Biozone"""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    
    # Find header row (contains 'Top depth')
    header_row = None
    data_start = None
    coords = None
    
    for i, row in enumerate(rows):
        row_str = [str(c).lower() if c else '' for c in row]
        if any('top depth' in s for s in row_str):
            header_row = i
            data_start = i + 1
            break
        # Check for coordinates
        try:
            if row[0] and row[2] and float(row[0]) > 100000:
                coords = (float(row[0]), float(row[2]))
        except (ValueError, TypeError, IndexError):
            pass
    
    if data_start is None:
        return None
    
    samples = []
    for row in rows[data_start:]:
        if row[0] is None:
            break
        try:
            md = float(row[0])
        except (ValueError, TypeError):
            continue
        
        zonation = int(row[1]) if row[1] is not None else None
        facies3 = int(row[2]) if row[2] is not None else None
        facies10 = int(row[3]) if row[3] is not None else None
        distality = int(row[4]) if row[4] is not None else None
        biozone = None
        if len(row) > 5 and row[5] is not None:
            try:
                bz = int(row[5])
                if bz != -999:
                    biozone = bz
            except (ValueError, TypeError):
                pass
        
        samples.append({
            'md': md,
            'zonation': zonation,
            'facies3': facies3,
            'facies10': facies10,
            'distality': distality,
            'biozone': biozone,
        })
    
    return {'coords': coords or WELL_COORDS.get(sheet_name) or APPROX_COORDS.get(sheet_name, (530000, 6750000)),
            'samples': samples}


def parse_format_b(ws, sheet_name):
    """Parse Format B sheets: Zonation, Top depth(MD), Dep.env. [Facies3 code or name]"""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    
    # Find header row
    header_row = None
    data_start = None
    
    for i, row in enumerate(rows):
        row_str = [str(c).lower() if c else '' for c in row]
        if any('top depth' in s or 'zonation' in s for s in row_str):
            header_row = i
            data_start = i + 1
            break
    
    if data_start is None:
        data_start = 2  # fallback
    
    samples = []
    current_zonation = None
    current_sequence = None
    
    # Map sequence names to numbers
    seq_map = {
        'series 6.2': 7, 'series 6.1': 6, 'series 6': 6,
        'series 5': 5, 'series 4': 4, 'series 3': 3, 'series 2': 2,
        'heather fm.': 8, 'heather c': 8,
        'sognefjord fm.': 7,
    }
    
    # Map dep env names to Facies10
    depenv_name_map = {
        'distributary channel': 1, 'dist. channel': 1,
        'foreshore': 2, 'tidal channel': 2,
        'tidal bar': 3, 'mixed tidal flat': 3, 'sandy tidal flat': 3,
        'subtidal flat': 4,
        'floodplain lake': 5, 'lake': 5, 'floodplain': 5,
        'mouth bar': 6,
        'upper delta front': 7, 'upper shoreface': 7,
        'lower delta front': 8, 'lower shoreface': 8,
        'prodelta': 9,
        'shelf': 10, 'offshore': 10, 'offshore transition': 10,
        'sandy debrite': 3, 'sandy debrite?': 3,
        'transgressive unit': 7,
    }
    
    for row in rows[data_start:]:
        # Zonation column (col 0)
        if row[0] is not None and str(row[0]).strip():
            zon_str = str(row[0]).strip().lower()
            for key, val in seq_map.items():
                if key in zon_str:
                    current_sequence = val
                    break
        
        # Top depth (col 1)
        md = None
        try:
            md = float(row[1]) if row[1] is not None else None
        except (ValueError, TypeError):
            continue
        
        if md is None:
            continue
        
        # Dep env (col 2) — could be Facies3 code or name string
        facies10 = None
        facies3 = None
        if row[2] is not None:
            dep_str = str(row[2]).strip()
            try:
                facies3 = int(dep_str)
                # It's a Facies3 code (1=distal, 2=intermediate, 3=proximal)
            except ValueError:
                # It's a name
                dep_lower = dep_str.lower()
                for key, val in depenv_name_map.items():
                    if key in dep_lower:
                        facies10 = val
                        break
                if facies10 is None and dep_str:
                    # Unknown name — try to infer from context
                    if '1' in dep_str:
                        facies3 = 1
                    elif '3' in dep_str and 'mudstone' not in dep_str.lower():
                        facies3 = 3
        
        # Infer distality from facies10 or facies3
        distality = None
        if facies10 is not None:
            distality = FACIES10_TO_DISTALITY.get(facies10)
            facies3 = 3 if facies10 <= 4 else (2 if facies10 <= 8 else 1)
        elif facies3 is not None:
            distality = {1: 1, 2: 2, 3: 4}.get(facies3)
        
        samples.append({
            'md': md,
            'zonation': current_sequence,
            'facies3': facies3,
            'facies10': facies10,
            'distality': distality,
            'biozone': None,  # Not available in Format B
        })
    
    coords = WELL_COORDS.get(sheet_name) or APPROX_COORDS.get(sheet_name, (530000, 6750000))
    return {'coords': coords, 'samples': samples}


def extract_all_wells():
    """Extract all wells from the Excel source."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    skip_sheets = {'FACIES', 'BIOZONE', 'SEQUENCE', 'Correlation', 
                   'Facies distribution', 'Legend', 'Sheet2', 'Sheet27'}
    
    # Format A sheets (have explicit Facies10 column)
    format_a_sheets = {'31_6-1', '31_6-5', '31_6-1 (2)', '31_6-5 (2)',
                       '31_6-2 (2)', '31_6-3 (2)'}
    
    wells = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue
        
        # Clean sheet name for well identifier
        well_name = sheet_name.replace('/', '_').replace(' ', '')
        # Skip duplicate sheets (prefer non-(2) versions for named wells)
        if '(2)' in sheet_name:
            base_name = sheet_name.replace(' (2)', '').replace('/', '_')
            if base_name in wells:
                continue
            well_name = base_name
        
        if sheet_name in format_a_sheets:
            result = parse_format_a(ws, sheet_name.replace(' (2)', '').replace('/', '_'))
        else:
            result = parse_format_b(ws, sheet_name.replace('/', '_'))
        
        if result and len(result['samples']) >= 3:
            wells[well_name] = result
    
    # Add biozone data from BIOZONE sheet
    ws_bio = wb['BIOZONE']
    for row in ws_bio.iter_rows(min_row=1, max_row=ws_bio.max_row, values_only=True):
        if row[0] is None:
            continue
        well_name = str(row[0]).replace('/', '_')
        if well_name not in wells:
            continue
        try:
            biozone_id = int(row[1])
            depth_top = float(row[2])
            depth_base = float(row[3])
        except (ValueError, TypeError, IndexError):
            continue
        # Assign biozone to samples in range
        for s in wells[well_name]['samples']:
            if depth_top <= s['md'] <= depth_base:
                s['biozone'] = biozone_id
    
    # Add sequence data from SEQUENCE sheet
    ws_seq = wb['SEQUENCE']
    for row in ws_seq.iter_rows(min_row=1, max_row=ws_seq.max_row, values_only=True):
        if row[0] is None:
            continue
        well_name = str(row[0]).replace('/', '_')
        if well_name not in wells:
            continue
        try:
            seq_id = int(row[1])
            depth_top = float(row[2])
            depth_base = float(row[3])
        except (ValueError, TypeError, IndexError):
            continue
        for s in wells[well_name]['samples']:
            if depth_top <= s['md'] <= depth_base:
                if s['zonation'] is None:
                    s['zonation'] = seq_id
    
    return wells


def write_weco_wells_txt(wells, filepath):
    """Write WeCo WellList format with all channels."""
    with open(filepath, 'w') as f:
        f.write("WeCo WellList 2\n")
        f.write(f"{len(wells)}\n")
        
        for well_name, data in sorted(wells.items()):
            samples = data['samples']
            if not samples:
                continue
            
            n = len(samples)
            x, y = data['coords']
            z_top = samples[0]['md']
            h = samples[-1]['md'] - samples[0]['md']
            
            f.write(f"{well_name}\n")
            f.write(f"{n}\n")
            f.write(f"{x:.6f} {y:.6f} {z_top:.6f} {h:.6f}\n")
            
            # Channels: MD, FACIES10, DISTALITY, BIOZONE, SEQUENCE
            num_channels = 5
            f.write(f"{num_channels}\n")
            
            # MD
            f.write(f"MD {n}\n")
            for s in samples:
                f.write(f"{s['md']:.6f}\n")
            
            # FACIES (use Facies10 where available, else Facies3)
            f.write(f"FACIES {n}\n")
            for s in samples:
                val = s['facies10'] if s['facies10'] is not None else (s['facies3'] or 0)
                f.write(f"{float(val):.6f}\n")
            
            # DISTALITY
            f.write(f"DISTALITY {n}\n")
            for s in samples:
                val = s['distality'] if s['distality'] is not None else 0
                f.write(f"{float(val):.6f}\n")
            
            # BIOZONE
            f.write(f"BIOZONE {n}\n")
            for s in samples:
                val = s['biozone'] if s['biozone'] is not None else 0
                f.write(f"{float(val):.6f}\n")
            
            # SEQUENCE
            f.write(f"SEQUENCE {n}\n")
            for s in samples:
                val = s['zonation'] if s['zonation'] is not None else 0
                f.write(f"{float(val):.6f}\n")
            
            # Regions (stratification)
            num_regions = 3
            f.write(f"{num_regions}\n")
            
            # BIOZONE regions
            bio_regions = _make_regions(samples, 'biozone')
            f.write(f"BIOZONE {len(bio_regions)}\n")
            for reg_id, start, length in bio_regions:
                f.write(f"{reg_id} {start} {length}\n")
            
            # FACIES regions  
            fac_regions = _make_regions_facies(samples)
            f.write(f"FACIES {len(fac_regions)}\n")
            for reg_id, start, length in fac_regions:
                f.write(f"{reg_id} {start} {length}\n")
            
            # SEQUENCE regions
            seq_regions = _make_regions(samples, 'zonation')
            f.write(f"SEQUENCE {len(seq_regions)}\n")
            for reg_id, start, length in seq_regions:
                f.write(f"{reg_id} {start} {length}\n")
        
        f.write("END\n")


def _make_regions(samples, key):
    """Create region intervals from sample data."""
    regions = []
    current_id = None
    start = 0
    length = 0
    
    for i, s in enumerate(samples):
        val = s[key]
        if val is None or val == 0:
            if current_id is not None:
                regions.append((current_id, start, length))
                current_id = None
            continue
        if val != current_id:
            if current_id is not None:
                regions.append((current_id, start, length))
            current_id = val
            start = i
            length = 1
        else:
            length += 1
    
    if current_id is not None:
        regions.append((current_id, start, length))
    return regions


def _make_regions_facies(samples):
    """Create facies region intervals."""
    regions = []
    current_id = None
    start = 0
    length = 0
    
    for i, s in enumerate(samples):
        val = s['facies10'] if s['facies10'] is not None else s.get('facies3')
        if val is None or val == 0:
            if current_id is not None:
                regions.append((current_id, start, length))
                current_id = None
            continue
        if val != current_id:
            if current_id is not None:
                regions.append((current_id, start, length))
            current_id = val
            start = i
            length = 1
        else:
            length += 1
    
    if current_id is not None:
        regions.append((current_id, start, length))
    return regions


def write_weco_json(wells, filepath):
    """Write wells.weco.json format."""
    well_list = []
    all_channels = ['MD', 'FACIES', 'DISTALITY', 'BIOZONE', 'SEQUENCE']
    region_names = ['BIOZONE', 'FACIES', 'SEQUENCE']
    
    for well_name, data in sorted(wells.items()):
        samples = data['samples']
        if not samples:
            continue
        
        n = len(samples)
        x, y = data['coords']
        z_top = samples[0]['md']
        h = samples[-1]['md'] - samples[0]['md']
        
        well_obj = {
            'name': well_name,
            'size': n,
            'location': {'x': x, 'y': y, 'z': z_top, 'h': h},
            'data': [
                {'name': 'MD', 'values': [s['md'] for s in samples]},
                {'name': 'FACIES', 'values': [float(s['facies10'] or s.get('facies3') or 0) for s in samples]},
                {'name': 'DISTALITY', 'values': [float(s['distality'] or 0) for s in samples]},
                {'name': 'BIOZONE', 'values': [float(s['biozone'] or 0) for s in samples]},
                {'name': 'SEQUENCE', 'values': [float(s['zonation'] or 0) for s in samples]},
            ],
            'regions': []
        }
        
        # Build region intervals
        for rname, key in [('BIOZONE', 'biozone'), ('SEQUENCE', 'zonation')]:
            intervals = []
            for reg_id, start, length in _make_regions(samples, key):
                intervals.append({'id': reg_id, 'start': start, 'length': length})
            well_obj['regions'].append({'name': rname, 'intervals': intervals})
        
        # Facies regions
        fac_intervals = []
        for reg_id, start, length in _make_regions_facies(samples):
            fac_intervals.append({'id': reg_id, 'start': start, 'length': length})
        well_obj['regions'].append({'name': 'FACIES', 'intervals': fac_intervals})
        
        well_list.append(well_obj)
    
    output = {
        'kind': 'weco:wbs:WellList:1.0.0',
        'meta': {
            'dataChannels': all_channels,
            'regionNames': region_names,
            'wellCount': len(well_list),
            'source': 'Troll field, Sognefjord Formation (Upper Jurassic)',
            'faciesDictionary': FACIES10_NAMES,
        },
        'wells': well_list
    }
    
    with open(filepath, 'w') as f:
        json.dump(output, f, indent=2)


def write_options(dirpath):
    """Write correlation option files for the Troll dataset."""
    
    # Option 1: Distal CCF (main approach for deltaic/shoreface)
    with open(os.path.join(dirpath, 'options_distal.txt'), 'w') as f:
        f.write("""# Config: TROLL — Distal CCF for Sognefjord Fm. deltaic sequences
# Primary: facies-based distality comparison (Walther's Law)
# No-crossing: BIOZONE (chronostratigraphic constraint)
# Wells represent proximal (distributary) to distal (prodelta/shelf) transect
#
cost-function=distal
dist-distal=DISTALITY
dist-facies=FACIES
dist-scaling=1.0
no-crossing=BIOZONE
order=position
max-cor=100
nbr-cor=50
out-nbr-cor=10
min-dist=0.2
out-min-dist=0.1
out-file=result_distal.txt
""")
    
    # Option 2: Variance + same-region (sequence constrained)
    with open(os.path.join(dirpath, 'options_sequence.txt'), 'w') as f:
        f.write("""# Config: TROLL — Variance CCF constrained by sequence boundaries
# Uses FACIES as correlation log with SEQUENCE as same-region constraint
# BIOZONE as no-crossing (hard chronostratigraphic tie)
#
cost-function=varsr
var-data=FACIES
same-region=SEQUENCE
no-crossing=BIOZONE
order=position
max-cor=100
nbr-cor=50
out-nbr-cor=10
min-dist=0.2
out-min-dist=0.1
out-file=result_sequence.txt
""")
    
    # Option 3: Basic variance (unconstrained for comparison)
    with open(os.path.join(dirpath, 'options_basic.txt'), 'w') as f:
        f.write("""# Config: TROLL — Basic variance CCF (unconstrained)
# Correlates FACIES without chronostratigraphic constraints
# Useful to show how unconstrained correlation can miscorrelate
#
cost-function=var
var-data=FACIES
order=position
max-cor=100
nbr-cor=50
out-nbr-cor=10
min-dist=0.3
out-min-dist=0.1
out-file=result_basic.txt
""")


def write_readme(dirpath, wells):
    """Write dataset README."""
    with open(os.path.join(dirpath, 'ReadMe.md'), 'w') as f:
        f.write(f"""# Troll Field Dataset

## Source
Sognefjord Formation (Upper Jurassic), Troll field, North Sea (blocks 31/2, 31/3, 31/5, 31/6).
Extracted from core-based sedimentological interpretation spreadsheet.

## Wells: {len(wells)}
{', '.join(sorted(wells.keys()))}

## Data Channels
- **MD**: Measured depth (m)
- **FACIES**: Depositional environment (Facies10 scheme, 1-10)
- **DISTALITY**: Proximal-distal position (1=distal/shelf ... 4=proximal/channel)
- **BIOZONE**: Biostratigraphic zonation (chronostratigraphic constraint)
- **SEQUENCE**: Depositional sequence (Series 2-7)

## Facies Legend (Facies10)
| Code | Environment | Distality |
|------|-------------|-----------|
| 1 | Distributary channel | 4 (proximal) |
| 2 | Foreshore / Tidal channel | 4 (proximal) |
| 3 | Tidal bar / Mixed flat | 3 (intermediate) |
| 4 | Subtidal flat | 3 (intermediate) |
| 5 | Floodplain lake | 2 (intermediate) |
| 6 | Mouth bar | 2 (intermediate) |
| 7 | Upper delta front | 2 (intermediate) |
| 8 | Lower delta front | 2 (distal) |
| 9 | Prodelta | 1 (distal) |
| 10 | Shelf / Offshore | 1 (distal) |

## Correlation Strategy
1. **options_distal.txt**: Distal CCF — uses facies + distality (Walther's Law)
2. **options_sequence.txt**: Variance + same-region (SEQUENCE boundaries)
3. **options_basic.txt**: Unconstrained variance (baseline comparison)

All constrained by BIOZONE no-crossing (where available).

## Geological Context
The Sognefjord Formation is a wave/tide-influenced deltaic system prograding
northward. Wells span proximal (distributary channels, foreshores) to distal
(prodelta, shelf) environments. Sequences 2-7 represent major regressive-
transgressive cycles bounded by flooding surfaces.
""")


if __name__ == '__main__':
    print("Extracting wells from Troll Excel source...")
    wells = extract_all_wells()
    print(f"  Found {len(wells)} wells")
    for name, data in sorted(wells.items()):
        n = len(data['samples'])
        has_bio = sum(1 for s in data['samples'] if s['biozone'])
        has_f10 = sum(1 for s in data['samples'] if s['facies10'])
        print(f"  {name:12s}: {n:3d} samples, {has_f10:2d} with Facies10, {has_bio:2d} with biozone")
    
    outdir = os.path.dirname(os.path.abspath(__file__))
    
    print(f"\nWriting wells.txt...")
    write_weco_wells_txt(wells, os.path.join(outdir, 'wells.txt'))
    
    print(f"Writing wells.weco.json...")
    write_weco_json(wells, os.path.join(outdir, 'wells.weco.json'))
    
    print(f"Writing option files...")
    write_options(outdir)
    
    print(f"Writing ReadMe.md...")
    write_readme(outdir, wells)
    
    print(f"\nDone! Troll dataset rebuilt with {len(wells)} wells.")
