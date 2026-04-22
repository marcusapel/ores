#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Strat Column Handler
===========================

Purpose
-------
A generic stratigraphic column model converter for **SMDA (.xlsx)**,
**RESQML 2.0.1 JSON graph**, and **OSDU Work-Product-Component (WPC) JSON bundle**.
It supports an **arbitrary number of ranks** per column, **mixed lithostratigraphy
& chronostratigraphy** in the same column, and **HorizonInterpretation** records
representing the boundaries (Top/Base) of stratigraphic units.

Supported conversions
---------------------
- SMDA (.xlsx, sheet *ApiStratUnit*) → RESQML JSON graph
- SMDA (.xlsx) → OSDU WPC bundle
- RESQML JSON graph → OSDU WPC bundle
- OSDU WPC bundle → RESQML JSON graph

CLI Usage (examples)
--------------------
SMDA → RESQML:
    python stratcolumnhandler.py smda2resqml --xlsx smda-api_strat-units.xlsx \
        --sheet ApiStratUnit -o smda.resqml.json

SMDA → OSDU:
    python stratcolumnhandler.py smda2osdu --xlsx smda-api_strat-units.xlsx \
        --sheet ApiStratUnit --partition data -o smda.osdu.json

RESQML → OSDU:
    python stratcolumnhandler.py resqml2osdu --resqml-json smda.resqml.json \
        --partition data -o smda_roundtrip.osdu.json

OSDU → RESQML:
    python stratcolumnhandler.py osdu2resqml --manifest smda.osdu.json \
        -o osdu2.resqml.json
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

try:
    import openpyxl  # only needed for .xlsx SMDA
except Exception:  # pragma: no cover
    openpyxl = None

try:
    import requests as _requests  # only needed for SMDA API fetch
except Exception:  # pragma: no cover
    _requests = None

# ------------------------------ Constants ---------------------------------
DEFAULT_PARTITION = os.getenv("OSDU_PARTITION", "data")  # default per OSDU examples
KIND_COL = "osdu:wks:work-product-component--StratigraphicColumn:1.2.0"
KIND_RANK = "osdu:wks:work-product-component--StratigraphicColumnRankInterpretation:1.3.0"
KIND_UNIT = "osdu:wks:work-product-component--StratigraphicUnitInterpretation:1.3.0"
KIND_HORIZON = "osdu:wks:work-product-component--HorizonInterpretation:1.2.0"
ORDERING_DEFAULT = "OlderToYounger"

# Generic vendor→OSDU mapping (paths relative to the **record root**).
# By default we map vendor fields into VendorMetadata.Raw to avoid claiming
# WKS semantics without explicit mapping.
VENDOR_TO_OSDU_MAP_DEFAULT = {
    # Common SMDA fields on units
    "source": "data.VendorMetadata.Raw.Source",
    "update_date": "data.VendorMetadata.Raw.UpdateDate",
    "update_user": "data.VendorMetadata.Raw.UpdateUser",
    "insert_date": "data.VendorMetadata.Raw.InsertDate",
    "insert_user": "data.VendorMetadata.Raw.InsertUser",
    "strat_unit_type": "data.VendorMetadata.Raw.UnitType",
    "color_html": "data.VendorMetadata.Raw.ColorHtml",
    "strat_unit_parent": "data.VendorMetadata.Raw.ParentName",
    "top_age": "data.VendorMetadata.Raw.TopAgeMa",
    "base_age": "data.VendorMetadata.Raw.BaseAgeMa",
}

# ------------------------------ Utilities ---------------------------------
def sanitize(name: str) -> str:
    """Return a filesystem/ID-safe token derived from a display name."""
    s = (name or "").strip()
    return "".join(ch if ch.isalnum() or ch in "-._" else "_" for ch in s) or "unnamed"


def wpc_id(partition: str, entity: str, name_or_uuid: str) -> str:
    """Construct an OSDU WPC record id."""
    return f"{partition}:work-product-component--{entity}:{sanitize(name_or_uuid)}:"


def load_reference_index(files: Optional[List[str]]) -> Dict[str, str]:
    """
    Build a lookup: lower(name/alias/code) -> record.id (SRN).

    Accepts either a bundle with "records":[], a list of records, or a single record.
    """
    if not files:
        return {}
    idx: Dict[str, str] = {}
    for path in files:
        with open(path, "r", encoding="utf-8") as f:
            doc = json.load(f)
        records = doc.get("records", doc if isinstance(doc, list) else [doc])
        for r in records:
            rid = r.get("id")
            data = r.get("data", {})
            if not rid:
                continue
            names: List[str] = []
            if data.get("Name"):
                names.append(str(data["Name"]))
            for a in (data.get("AliasNames") or []):
                if a:
                    names.append(str(a))
            for k in ("Code", "CodeAsNumber"):
                if data.get(k) is not None:
                    names.append(str(data[k]))
            for val in names:
                key = val.lower().strip()
                if key:
                    idx[key] = rid
    return idx


def _set_path(root: Dict[str, Any], path: str, value: Any) -> None:
    """Set a dotted path into a dict, creating intermediate dicts as needed."""
    parts = path.split(".")
    cur: Any = root
    for p in parts[:-1]:
        if p not in cur or not isinstance(cur[p], dict):
            cur[p] = {}
        cur = cur[p]
    cur[parts[-1]] = value


def apply_vendor_map(src: Dict[str, Any], dest_record: Dict[str, Any], mapping: Dict[str, str]) -> None:
    """Apply a mapping from vendor keys to dotted OSDU paths in the record."""
    for skey, tpath in mapping.items():
        if skey in src and src[skey] is not None:
            _set_path(dest_record, tpath, src[skey])


# -------------------------------- Model -----------------------------------
@dataclass
class StratUnit:
    name: str
    uuid: str = field(default_factory=lambda: str(uuid.uuid4()))
    level: Optional[int] = None
    top_age_ma: Optional[float] = None
    base_age_ma: Optional[float] = None
    parent_name: Optional[str] = None
    color_html: Optional[str] = None
    vendor: Dict[str, Any] = field(default_factory=dict)  # raw vendor row fields


@dataclass
class StratHorizon:
    """Boundary between two stratigraphic units (Top/Base of a unit)."""
    name: str
    uuid: str = field(default_factory=lambda: str(uuid.uuid4()))
    unit_name: Optional[str] = None          # unit this horizon bounds
    unit_id: Optional[str] = None            # OSDU record id of the unit
    boundary_type: str = "Top"               # 'Top' | 'Base'
    age_ma: Optional[float] = None
    vendor: Dict[str, Any] = field(default_factory=dict)


@dataclass
class StratRank:
    name: str
    kind: str  # 'litho' | 'chrono'
    level: Optional[int] = None
    ordering: str = ORDERING_DEFAULT
    units: List[StratUnit] = field(default_factory=list)
    chrono_names: List[str] = field(default_factory=list)  # names or SRNs
    vendor: Dict[str, Any] = field(default_factory=dict)


@dataclass
class StratColumn:
    name: str
    ranks: List[StratRank]
    horizons: List[StratHorizon] = field(default_factory=list)
    vendor: Dict[str, Any] = field(default_factory=dict)

    # ------------------------------- Importers -----------------------------
    @staticmethod
    def from_smda_xlsx(xlsx_path: str, sheet: str = "ApiStratUnit") -> "StratColumn":
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to read .xlsx files")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {xlsx_path}")
        ws = wb[sheet]

        headers = [
            str(h).strip() if h is not None else ""
            for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        h2i = {h: i for i, h in enumerate(headers)}

        def cell(row, key):
            i = h2i.get(key)
            return None if i is None else row[i]

        def _float(x):
            try:
                return float(x)
            except Exception:
                return None

        def _int(x):
            try:
                return int(float(x))
            except Exception:
                return None

        items: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue

            d = {h: row[idx] if idx < len(row) else None for h, idx in h2i.items()}
            name = (cell(row, "identifier") or "").strip()
            level = _int(cell(row, "strat_unit_level")) or 9999
            parent = cell(row, "strat_unit_parent")
            if isinstance(parent, str) and parent.lower() in ("", "null", "none"):
                parent = None

            items.append(
                {
                    **d,
                    "__name": name,
                    "__level": level,
                    "__parent": parent,
                    "__top": _float(cell(row, "top_age")),
                    "__base": _float(cell(row, "base_age")),
                    "__color": cell(row, "color_html") or None,
                    "__coltype": str(cell(row, "strat_column_type") or "").lower(),
                    "__uuid": str(cell(row, "uuid") or uuid.uuid4()),
                    "__colname": cell(row, "strat_column_identifier") or "Stratigraphic Column",
                }
            )

        wb.close()

        if not items:
            raise ValueError("No data rows in SMDA sheet")

        col_name = next((it["__colname"] for it in items if it["__colname"]), "Stratigraphic Column")

        by_level: Dict[int, List[Dict[str, Any]]] = {}
        for it in items:
            by_level.setdefault(it["__level"], []).append(it)

        ranks: List[StratRank] = []
        for lvl in sorted(by_level.keys()):
            rows = by_level[lvl]
            is_chrono = any("chronostrat" in (it["__coltype"] or "") for it in rows)

            def age_key(x):
                ta = x["__top"] if x["__top"] is not None else 1e12
                ba = x["__base"] if x["__base"] is not None else 1e12
                nm = (x["__name"] or "").lower()
                return (ta, ba, nm)

            rows = sorted(rows, key=age_key)
            label = f"Level{lvl}"

            if is_chrono:
                ranks.append(
                    StratRank(
                        name=label,
                        kind="chrono",
                        level=lvl,
                        chrono_names=[(r["__name"] or "").strip() for r in rows if r["__name"]],
                    )
                )
            else:
                units: List[StratUnit] = []
                for r in rows:
                    vendor = {k: v for k, v in r.items() if not k.startswith("__")}
                    units.append(
                        StratUnit(
                            name=r["__name"] or "Unit",
                            uuid=r["__uuid"],
                            level=lvl,
                            top_age_ma=r["__top"],
                            base_age_ma=r["__base"],
                            parent_name=r["__parent"],
                            color_html=r["__color"],
                            vendor=vendor,
                        )
                    )
                ranks.append(StratRank(name=label, kind="litho", level=lvl, units=units))

        return StratColumn(name=col_name, ranks=ranks, vendor={})

    # ---- SMDA column-header importers (ApiStratColumn) ----

    # 15 canonical SMDA column-header fields
    SMDA_COLUMN_FIELDS = [
        "identifier", "strat_column_type", "strat_column_status",
        "strat_column_area_type", "country", "area", "field", "source",
        "update_date", "update_user", "insert_date", "insert_user",
        "description", "interpreter", "uuid",
    ]

    @staticmethod
    def from_smda_column_csv(
        csv_path: str,
        delimiter: str = ";",
    ) -> List["StratColumn"]:
        """Import SMDA column headers from a CSV export.

        Each row becomes one StratColumn (no ranks/units - just header metadata).
        All 15 SMDA fields are preserved in StratColumn.vendor.

        Returns
        -------
        list[StratColumn]
            One StratColumn per row, with vendor dict containing all fields.
        """
        import csv as _csv

        with open(csv_path, "r", encoding="utf-8") as f:
            reader = _csv.DictReader(f, delimiter=delimiter)
            rows = list(reader)

        columns: List[StratColumn] = []
        for row in rows:
            name = (row.get("identifier") or "").strip()
            if not name:
                continue
            vendor = {}
            for k in StratColumn.SMDA_COLUMN_FIELDS:
                v = row.get(k)
                # Preserve original value exactly (including trailing spaces)
                # Only normalise truly empty strings to None
                if v is None or v == "":
                    vendor[k] = None
                else:
                    vendor[k] = v
            columns.append(StratColumn(name=name, ranks=[], vendor=vendor))
        return columns

    @staticmethod
    def from_smda_column_xlsx(
        xlsx_path: str,
        sheet: str = "ApiStratColumn",
    ) -> List["StratColumn"]:
        """Import SMDA column headers from an XLSX export.

        Each row becomes one StratColumn (no ranks/units - just header metadata).
        All 15 SMDA fields are preserved in StratColumn.vendor.

        Returns
        -------
        list[StratColumn]
            One StratColumn per row, with vendor dict containing all fields.
        """
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to read .xlsx files")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {xlsx_path}")
        ws = wb[sheet]

        headers = [
            str(h).strip() if h is not None else ""
            for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        h2i = {h: i for i, h in enumerate(headers)}

        columns: List[StratColumn] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            name_idx = h2i.get("identifier")
            name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ""
            if not name:
                continue

            vendor: Dict[str, Any] = {}
            for fld in StratColumn.SMDA_COLUMN_FIELDS:
                idx = h2i.get(fld)
                val = row[idx] if idx is not None and idx < len(row) else None
                # Preserve original value exactly (convert to str for consistency)
                if val is None:
                    vendor[fld] = None
                else:
                    s = str(val)
                    vendor[fld] = s if s else None
            columns.append(StratColumn(name=name, ranks=[], vendor=vendor))
        wb.close()
        return columns

    # ---- SMDA column-header exporter ----

    def to_smda_column_dict(self) -> Dict[str, Any]:
        """Export this StratColumn's metadata back to an SMDA column-header dict.

        Returns a dict with the 15 canonical SMDA column fields, reconstructed
        from vendor metadata.  This is the inverse of from_smda_column_csv /
        from_smda_column_xlsx.
        """
        out: Dict[str, Any] = {}
        v = self.vendor or {}
        for fld in StratColumn.SMDA_COLUMN_FIELDS:
            out[fld] = v.get(fld)
        # Ensure name is always populated from self.name
        if not out.get("identifier"):
            out["identifier"] = self.name
        return out

    # ---- OpenWorks JSON importer ----

    @staticmethod
    def from_openworks_json(path_or_obj) -> "StratColumn":
        """Import a stratigraphic column from OpenWorks JSON.

        Expected schema (flexible):
        {
          "StratColumn": {"Name": "...", "Type": "litho|chrono|mixed"},
          "Units": [
            {
              "Name": "Draupne Fm",
              "Level": 2,              // rank grouping: 1=Group, 2=Formation, …
              "ParentName": "Viking Gp",
              "TopAge": 149.2,
              "BaseAge": 163.5,
              "Color": "#AABB00",
              "UnitType": "lithostratigraphic",
              "UUID": "...",           // optional stable id
              ... any extra vendor fields
            },
            ...
          ]
        }

        Also accepts a list of units (flat): [{...}, ...]
        or a dict with "units" (lowercase): {"units": [...]}
        """
        if isinstance(path_or_obj, (str, os.PathLike)):
            with open(path_or_obj, "r", encoding="utf-8") as f:
                doc = json.load(f)
        else:
            doc = path_or_obj  # already a dict/list

        # Normalize shape
        if isinstance(doc, list):
            doc = {"StratColumn": {"Name": "OpenWorks Column"}, "Units": doc}
        if "units" in doc and "Units" not in doc:
            doc["Units"] = doc.pop("units")
        if "StratColumn" not in doc and "stratColumn" in doc:
            doc["StratColumn"] = doc.pop("stratColumn")
        if "StratColumn" not in doc:
            doc["StratColumn"] = {"Name": "OpenWorks Column"}

        sc = doc.get("StratColumn") or {}
        raw_units = doc.get("Units") or []
        if not raw_units:
            raise ValueError("OpenWorks JSON must contain 'Units' (array of unit objects)")

        col_name = sc.get("Name") or "OpenWorks Column"
        col_type = (sc.get("Type") or "litho").lower()

        def _float(x):
            try:
                return float(x)
            except (TypeError, ValueError):
                return None

        def _int(x):
            try:
                return int(float(x))
            except (TypeError, ValueError):
                return None

        # Normalize units with flexible field names
        items = []
        for u in raw_units:
            if not isinstance(u, dict):
                continue
            name = (u.get("Name") or u.get("name") or u.get("identifier") or "").strip()
            if not name:
                continue

            level = _int(
                u.get("Level") or u.get("level") or u.get("strat_unit_level") or 0
            )
            parent = (
                u.get("ParentName") or u.get("parentName")
                or u.get("parent_name") or u.get("strat_unit_parent")
            )
            if isinstance(parent, str) and parent.lower() in ("", "null", "none"):
                parent = None
            top = _float(
                u.get("TopAge") or u.get("topAge") or u.get("top_age")
                or u.get("TopAgeMa") or u.get("OlderAge")
            )
            base = _float(
                u.get("BaseAge") or u.get("baseAge") or u.get("base_age")
                or u.get("BaseAgeMa") or u.get("YoungerAge")
            )
            color = (
                u.get("Color") or u.get("color") or u.get("color_html")
                or u.get("ColorHtml") or u.get("Colour")
            )
            unit_type = (
                u.get("UnitType") or u.get("unitType") or u.get("strat_unit_type")
                or u.get("strat_column_type") or ""
            ).lower()
            uid = str(u.get("UUID") or u.get("uuid") or u.get("Id") or uuid.uuid4())

            # All original fields as vendor data (exclude normalized ones)
            vendor = dict(u)

            items.append({
                "name": name,
                "level": level or 0,
                "parent": parent,
                "top": top,
                "base": base,
                "color": color,
                "type": unit_type,
                "uuid": uid,
                "vendor": vendor,
            })

        if not items:
            raise ValueError("No valid units found in OpenWorks JSON")

        # Group by level → ranks
        by_level: Dict[int, list] = {}
        for it in items:
            by_level.setdefault(it["level"], []).append(it)

        ranks: List[StratRank] = []
        for lvl in sorted(by_level.keys()):
            rows = by_level[lvl]
            is_chrono = any("chrono" in (r["type"] or "") for r in rows)

            rows.sort(key=lambda x: (
                x["top"] if x["top"] is not None else 1e12,
                x["base"] if x["base"] is not None else 1e12,
                x["name"].lower(),
            ))

            # Derive rank label from the highest-populated unit type or level
            label = f"Level{lvl}"
            if is_chrono:
                ranks.append(StratRank(
                    name=label, kind="chrono", level=lvl,
                    chrono_names=[r["name"] for r in rows if r["name"]],
                ))
            else:
                units = []
                for r in rows:
                    units.append(StratUnit(
                        name=r["name"],
                        uuid=r["uuid"],
                        level=lvl,
                        top_age_ma=r["top"],
                        base_age_ma=r["base"],
                        parent_name=r["parent"],
                        color_html=r["color"],
                        vendor=r["vendor"],
                    ))
                ranks.append(StratRank(name=label, kind="litho", level=lvl, units=units))

        return StratColumn(name=col_name, ranks=ranks, vendor={"source": "OpenWorks", "type": col_type})

    # ---- SMDA API importer ----

    @staticmethod
    def from_smda_api(
        column_identifier: str,
        *,
        base_url: str = "https://api.gateway.equinor.com",
        access_token: Optional[str] = None,
        api_key: Optional[str] = None,
        verify_ssl: bool = True,
    ) -> "StratColumn":
        """Fetch a stratigraphic column from the SMDA REST API.

        Supports both the Equinor API Gateway (api.gateway.equinor.com) and the
        legacy OPUS endpoint (opus.smda.equinor.com).  The endpoint path is
        chosen automatically based on the base_url.

        Parameters
        ----------
        column_identifier : str
            The strat_column_identifier to query (e.g. "NCS Lithostratigraphy",
            "DIAPIRIC PROVINCE LITHOSTRATIGRAPHY").
        base_url : str
            SMDA base URL (default: https://api.gateway.equinor.com).
        access_token : str | None
            Bearer token for Azure AD auth (Equinor SSO).
        api_key : str | None
            Ocp-Apim-Subscription-Key for API gateway access.
        verify_ssl : bool
            Whether to verify TLS certificates.

        Returns
        -------
        StratColumn
            Parsed column with ranks grouped by strat_unit_level.
        """
        if _requests is None:
            raise RuntimeError("'requests' package is required for SMDA API access. "
                               "Install with: pip install requests")

        headers: Dict[str, str] = {
            "Accept": "application/json",
            "Cache-Control": "no-cache",
        }
        if access_token:
            headers["Authorization"] = f"Bearer {access_token}"
        if api_key:
            headers["Ocp-Apim-Subscription-Key"] = api_key

        # Choose endpoint path based on base URL
        is_gateway = "api.gateway" in base_url.lower()
        base = base_url.rstrip("/")

        if is_gateway:
            # Equinor API Gateway: paginated REST endpoint
            url = f"{base}/smda/v2.0/smda-api/strat-units"
            all_rows: list = []
            page = 1
            while True:
                params = {
                    "_page": str(page),
                    "_items": "500",
                    "_order": "asc",
                    "strat_column_identifier": column_identifier,
                }
                resp = _requests.get(url, params=params, headers=headers,
                                     verify=verify_ssl, timeout=60)
                resp.raise_for_status()
                ct = (resp.headers.get("Content-Type") or "").lower()
                if "json" not in ct:
                    snippet = resp.text[:300].strip()
                    raise ValueError(
                        f"SMDA returned non-JSON response (Content-Type: {ct}). "
                        f"Response preview: {snippet}")
                data = resp.json()
                # Gateway returns {"data": {"results": [...], "hits": N, ...}}
                inner = data.get("data", data) if isinstance(data, dict) else data
                if isinstance(inner, dict):
                    rows = inner.get("results", inner.get("value", []))
                elif isinstance(inner, list):
                    rows = inner
                else:
                    rows = []
                if not isinstance(rows, list):
                    rows = []
                all_rows.extend(rows)
                # Pagination check
                total = None
                if isinstance(inner, dict):
                    total = inner.get("hits", inner.get("total",
                            inner.get("totalCount", inner.get("_total"))))
                if total is not None:
                    try:
                        total = int(total)
                    except (TypeError, ValueError):
                        total = None
                if total is not None and len(all_rows) >= total:
                    break
                if len(rows) < 500:
                    break
                page += 1
                if page > 20:
                    break
            rows = all_rows
        else:
            # Legacy OPUS: OData-style endpoint
            url = f"{base}/smda/table/v2/api_strat_unit_header"
            params = {
                "$filter": f"strat_column_identifier eq '{column_identifier}'",
                "$format": "json",
                "$top": "5000",
            }
            resp = _requests.get(url, params=params, headers=headers,
                                 verify=verify_ssl, timeout=60)
            resp.raise_for_status()
            ct = (resp.headers.get("Content-Type") or "").lower()
            if "json" not in ct:
                snippet = resp.text[:300].strip()
                raise ValueError(
                    f"SMDA returned non-JSON response (Content-Type: {ct}). "
                    f"Response preview: {snippet}")
            data = resp.json()
            rows = data.get("value") if isinstance(data, dict) else data
            if not isinstance(rows, list):
                rows = []

        if not rows:
            raise ValueError(
                f"No strat units returned from SMDA for column "
                f"'{column_identifier}'.  The column header exists in SMDA but "
                f"has no units in the strat-units table.  This is an SMDA data "
                f"gap - the column may not have been populated yet."
            )

        def _float(x):
            try:
                return float(x)
            except (TypeError, ValueError):
                return None

        def _int(x):
            try:
                return int(float(x))
            except (TypeError, ValueError):
                return None

        # Parse rows into items
        items = []
        for row in rows:
            name = str(row.get("identifier") or row.get("strat_unit_identifier") or "").strip()
            if not name:
                continue
            level = _int(row.get("strat_unit_level"))
            parent = row.get("strat_unit_parent")
            if isinstance(parent, str) and parent.lower() in ("", "null", "none"):
                parent = None
            top = _float(row.get("top_age"))
            base = _float(row.get("base_age"))
            color = row.get("color_html")
            col_type = str(row.get("strat_column_type") or "").lower()
            uid = str(row.get("uuid") or uuid.uuid4())
            col_name_raw = row.get("strat_column_identifier") or column_identifier

            items.append({
                "name": name,
                "level": level or 0,
                "parent": parent,
                "top": top,
                "base": base,
                "color": color,
                "type": col_type,
                "uuid": uid,
                "col_name": col_name_raw,
                "vendor": dict(row),
            })

        if not items:
            raise ValueError(f"No valid strat units parsed from SMDA response "
                             f"for column '{column_identifier}'")

        col_name = items[0]["col_name"]

        # Group by level → ranks
        by_level: Dict[int, list] = {}
        for it in items:
            by_level.setdefault(it["level"], []).append(it)

        ranks: List[StratRank] = []
        for lvl in sorted(by_level.keys()):
            level_rows = by_level[lvl]
            is_chrono = any("chronostrat" in (r["type"] or "") for r in level_rows)

            level_rows.sort(key=lambda x: (
                x["top"] if x["top"] is not None else 1e12,
                x["base"] if x["base"] is not None else 1e12,
                x["name"].lower(),
            ))

            label = f"Level{lvl}"
            if is_chrono:
                ranks.append(StratRank(
                    name=label, kind="chrono", level=lvl,
                    chrono_names=[r["name"] for r in level_rows if r["name"]],
                ))
            else:
                units = []
                for r in level_rows:
                    units.append(StratUnit(
                        name=r["name"],
                        uuid=r["uuid"],
                        level=lvl,
                        top_age_ma=r["top"],
                        base_age_ma=r["base"],
                        parent_name=r["parent"],
                        color_html=r["color"],
                        vendor=r["vendor"],
                    ))
                ranks.append(StratRank(name=label, kind="litho", level=lvl, units=units))

        return StratColumn(name=col_name, ranks=ranks,
                           vendor={"source": "SMDA", "api_url": url})

    @staticmethod
    def from_smda_api_list_columns(
        *,
        base_url: str = "https://api.gateway.equinor.com",
        access_token: Optional[str] = None,
        api_key: Optional[str] = None,
        verify_ssl: bool = True,
    ) -> List[str]:
        """List available stratigraphic column identifiers from SMDA.

        Returns a sorted list of distinct strat_column_identifier values.
        Supports both the Equinor API Gateway and the legacy OPUS endpoint.
        """
        if _requests is None:
            raise RuntimeError("'requests' package is required")

        headers: Dict[str, str] = {
            "Accept": "application/json",
            "Cache-Control": "no-cache",
        }
        if access_token:
            headers["Authorization"] = f"Bearer {access_token}"
        if api_key:
            headers["Ocp-Apim-Subscription-Key"] = api_key

        is_gateway = "api.gateway" in base_url.lower()
        base = base_url.rstrip("/")

        if is_gateway:
            url = f"{base}/smda/v2.0/smda-api/strat-column"
            all_rows: list = []
            page = 1
            while True:
                params = {
                    "_page": str(page),
                    "_items": "100",
                    "_order": "asc",
                    "_aggregation_include_buckets": "true",
                }
                resp = _requests.get(url, params=params, headers=headers,
                                     verify=verify_ssl, timeout=60)
                resp.raise_for_status()
                ct = (resp.headers.get("Content-Type") or "").lower()
                if "json" not in ct:
                    snippet = resp.text[:300].strip()
                    raise ValueError(
                        f"SMDA returned non-JSON (Content-Type: {ct}). Preview: {snippet}")
                data = resp.json()
                # Gateway returns {"data": {"results": [...], "hits": N, ...}}
                inner = data.get("data", data) if isinstance(data, dict) else data
                if isinstance(inner, dict):
                    rows = inner.get("results", inner.get("value", []))
                elif isinstance(inner, list):
                    rows = inner
                else:
                    rows = []
                if not isinstance(rows, list):
                    rows = []
                all_rows.extend(rows)
                total = None
                if isinstance(inner, dict):
                    total = inner.get("hits", inner.get("total",
                            inner.get("totalCount", inner.get("_total"))))
                if total is not None:
                    try:
                        total = int(total)
                    except (TypeError, ValueError):
                        total = None
                if total is not None and len(all_rows) >= total:
                    break
                if len(rows) < 100:
                    break
                page += 1
                if page > 50:
                    break
            names = sorted(set(
                str(r.get("strat_column_identifier") or r.get("identifier", "")).strip()
                for r in all_rows
                if (r.get("strat_column_identifier") or r.get("identifier", "")).strip()
            ))
        else:
            url = f"{base}/smda/table/v2/api_strat_unit_header"
            params = {
                "$select": "strat_column_identifier",
                "$format": "json",
                "$top": "10000",
            }
            resp = _requests.get(url, params=params, headers=headers,
                                 verify=verify_ssl, timeout=60)
            resp.raise_for_status()
            ct = (resp.headers.get("Content-Type") or "").lower()
            if "json" not in ct:
                snippet = resp.text[:300].strip()
                raise ValueError(
                    f"SMDA returned non-JSON (Content-Type: {ct}). Preview: {snippet}")
            data = resp.json()
            rows = data.get("value") if isinstance(data, dict) else data
            if not isinstance(rows, list):
                rows = []
            names = sorted(set(
                str(r.get("strat_column_identifier") or "").strip()
                for r in rows if r.get("strat_column_identifier")
            ))

        return names

    @staticmethod
    def from_resqml_json(path: str) -> "StratColumn":
        with open(path, "r", encoding="utf-8") as f:
            objs = json.load(f)
        objs = objs if isinstance(objs, list) else [objs]

        units = [o for o in objs if o.get("resqmlType") == "resqml20:StratigraphicUnitInterpretation"]
        ranks = [o for o in objs if o.get("resqmlType") == "resqml20:StratigraphicColumnRankInterpretation"]
        cols = [o for o in objs if o.get("resqmlType") == "resqml20:StratigraphicColumn"]
        horizon_objs = [o for o in objs if o.get("resqmlType") == "resqml20:HorizonInterpretation"]

        uid2u: Dict[str, StratUnit] = {}
        for u in units:
            uid = u.get("uuid") or str(uuid.uuid4())
            extra = u.get("extraMetadata", {}) or {}
            vendor_raw = extra.get("vendor", extra.get("ow", {}))  # accept legacy 'ow'
            uid2u[uid] = StratUnit(
                name=u.get("title") or "Unit",
                uuid=uid,
                level=extra.get("level"),
                top_age_ma=u.get("topAgeMa"),
                base_age_ma=u.get("baseAgeMa"),
                parent_name=u.get("parentName"),
                color_html=u.get("colorHtml"),
                vendor=vendor_raw,
            )

        sranks: List[StratRank] = []
        for r in ranks:
            name = r.get("title") or "Rank"
            ordering = r.get("orderingCriteria") or ORDERING_DEFAULT
            if r.get("unitInterpretationRefs"):
                us = [
                    uid2u.get(ref.get("uuid"))
                    or StratUnit(name=ref.get("uuid", "unit"), uuid=ref.get("uuid", ""))
                    for ref in r["unitInterpretationRefs"]
                ]
                sranks.append(StratRank(name=name, kind="litho", ordering=ordering, units=us))
            elif r.get("chronoStratRefs"):
                srns = [ref.get("srn") for ref in r["chronoStratRefs"]]
                sranks.append(StratRank(name=name, kind="chrono", ordering=ordering, chrono_names=srns))
            else:
                sranks.append(StratRank(name=name, kind="litho", ordering=ordering, units=[]))

        cname = cols[0].get("title") if cols else "Stratigraphic Column"
        col_vendor: Dict[str, Any] = {}
        if cols:
            cem = cols[0].get("extraMetadata") or {}
            col_vendor = cem.get("vendor", {})

        # Horizons
        shorizons: List[StratHorizon] = []
        for h in horizon_objs:
            uid = h.get("uuid") or str(uuid.uuid4())
            extra = h.get("extraMetadata", {}) or {}
            shorizons.append(StratHorizon(
                name=h.get("title") or "Horizon",
                uuid=uid,
                unit_name=extra.get("unitName"),
                boundary_type=extra.get("boundaryType", "Top"),
                age_ma=h.get("ageMa"),
                vendor=extra.get("vendor", {}),
            ))

        return StratColumn(name=cname or "Stratigraphic Column", ranks=sranks, horizons=shorizons, vendor=col_vendor)

    @staticmethod
    def from_osdu_bundle(path: str) -> "StratColumn":
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if not isinstance(payload, dict) or "records" not in payload:
            raise ValueError("OSDU payload must be a dict with 'records'")
        recs = payload["records"]

        _pfx_unit    = KIND_UNIT.rsplit(":", 1)[0] + ":"
        _pfx_rank    = KIND_RANK.rsplit(":", 1)[0] + ":"
        _pfx_col     = KIND_COL.rsplit(":", 1)[0] + ":"
        _pfx_horizon = KIND_HORIZON.rsplit(":", 1)[0] + ":"
        units    = [r for r in recs if (r.get("kind") or "").startswith(_pfx_unit)]
        ranks    = [r for r in recs if (r.get("kind") or "").startswith(_pfx_rank)]
        cols     = [r for r in recs if (r.get("kind") or "").startswith(_pfx_col)]
        horizons = [r for r in recs if (r.get("kind") or "").startswith(_pfx_horizon)]

        id2u: Dict[str, StratUnit] = {}
        for ur in units:
            d = ur.get("data", {})
            nm = d.get("Name") or "Unit"
            rid = ur.get("id") or ""
            uid = rid.split(":")[-2] if rid.endswith(":") else rid
            vm = (d.get("VendorMetadata", {}) or {})
            vendor_raw = vm.get("Raw", vm.get("OW", {}))  # accept legacy 'OW'
            # Read age, parent, color from structured fields OR vendor metadata
            tr = d.get("TimeRange", {}) or {}
            top_age = tr.get("TopAgeMa") or vendor_raw.get("TopAgeMa") or vendor_raw.get("top_age")
            base_age = tr.get("BaseAgeMa") or vendor_raw.get("BaseAgeMa") or vendor_raw.get("base_age")
            rel = d.get("Relationships", {}) or {}
            parent_name = (d.get("ParentName")
                           or (rel.get("Parent", {}) or {}).get("Name")
                           or vendor_raw.get("ParentName")
                           or vendor_raw.get("strat_unit_parent"))
            color_html = ((d.get("Rendering", {}) or {}).get("ColorHtml")
                          or vendor_raw.get("ColorHtml")
                          or vendor_raw.get("color_html"))
            try:
                top_age = float(top_age) if top_age is not None else None
            except (ValueError, TypeError):
                top_age = None
            try:
                base_age = float(base_age) if base_age is not None else None
            except (ValueError, TypeError):
                base_age = None
            id2u[ur.get("id")] = StratUnit(
                name=nm,
                uuid=uid or str(uuid.uuid4()),
                top_age_ma=top_age,
                base_age_ma=base_age,
                parent_name=parent_name,
                color_html=color_html,
                vendor=vendor_raw,
            )

        sranks: List[StratRank] = []
        for rr in ranks:
            d = rr.get("data", {})
            nm = d.get("Name") or "Rank"
            ordering = d.get("OrderingCriteria") or ORDERING_DEFAULT
            if d.get("StratigraphicUnitInterpretationSet"):
                seen_uids = set()
                us = []
                for x in d["StratigraphicUnitInterpretationSet"]:
                    if x in seen_uids:
                        continue
                    seen_uids.add(x)
                    us.append(id2u.get(x) or StratUnit(name=x or "unit", uuid=str(uuid.uuid4())))
                sranks.append(StratRank(name=nm, kind="litho", ordering=ordering, units=us))
            elif d.get("ChronoStratigraphySet"):
                srns = list(d["ChronoStratigraphySet"])
                sranks.append(StratRank(name=nm, kind="chrono", ordering=ordering, chrono_names=srns))
            else:
                sranks.append(StratRank(name=nm, kind="litho", ordering=ordering, units=[]))

        # Horizons
        shorizons: List[StratHorizon] = []
        for hr in horizons:
            d = hr.get("data", {})
            nm = d.get("Name") or "Horizon"
            rid = hr.get("id") or ""
            uid = rid.split(":")[-2] if rid.endswith(":") else rid
            vm = (d.get("VendorMetadata", {}) or {}).get("Raw", {})
            # Parse boundary type and linked unit
            btype = vm.get("BoundaryType", "Top")
            unit_name = vm.get("UnitName")
            unit_id_ref = None
            for ref in (d.get("InterpretedFeatureReferences") or []):
                if isinstance(ref, str):
                    unit_id_ref = ref
                    break
            age = None
            try:
                age = float(d.get("AgeMa") or vm.get("AgeMa"))
            except (ValueError, TypeError):
                pass
            shorizons.append(StratHorizon(
                name=nm, uuid=uid or str(uuid.uuid4()),
                unit_name=unit_name, unit_id=unit_id_ref,
                boundary_type=btype, age_ma=age, vendor=vm,
            ))

        cname = cols[0].get("data", {}).get("Name") if cols else "Stratigraphic Column"
        col_vendor: Dict[str, Any] = {}
        if cols:
            cvm = (cols[0].get("data", {}).get("VendorMetadata") or {})
            col_vendor = cvm.get("Raw", {})
        return StratColumn(name=cname or "Stratigraphic Column", ranks=sranks, horizons=shorizons, vendor=col_vendor)

    # ------------------------------- Exporters -----------------------------
    def to_resqml_json(self, chrono_rd_index: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
        objs: List[Dict[str, Any]] = []
        rank_objs: List[Dict[str, Any]] = []

        # Units (from litho ranks)
        for r in self.ranks:
            if r.kind != "litho":
                continue
            for u in r.units:
                _em = {"vendor": dict(u.vendor)}
                if u.level is not None:
                    _em["level"] = u.level
                obj = {
                    "resqmlType": "resqml20:StratigraphicUnitInterpretation",
                    "uuid": u.uuid,
                    "title": u.name,
                    "extraMetadata": _em,
                }
                if u.top_age_ma is not None:
                    obj["topAgeMa"] = u.top_age_ma
                if u.base_age_ma is not None:
                    obj["baseAgeMa"] = u.base_age_ma
                if u.parent_name:
                    obj["parentName"] = u.parent_name
                if u.color_html:
                    obj["colorHtml"] = u.color_html
                objs.append(obj)

        # Ranks (litho + chrono)
        for r in self.ranks:
            rid = sanitize(r.name)
            if r.kind == "litho":
                refs = [
                    {"uuid": u.uuid, "contentType": "resqml20:StratigraphicUnitInterpretation"} for u in r.units
                ]
                robj = {
                    "resqmlType": "resqml20:StratigraphicColumnRankInterpretation",
                    "uuid": rid,
                    "title": r.name,
                    "orderingCriteria": r.ordering,
                    "unitInterpretationRefs": refs,
                }
            else:
                srns: List[str] = []
                for nm in r.chrono_names:
                    s = (nm or "").strip()
                    if ":" in s and s.count(":") >= 2:
                        srns.append(s)
                    else:
                        if not chrono_rd_index:
                            raise ValueError(
                                f"Chrono name '{nm}' requires --chrono-refdata to resolve to SRN"
                            )
                        srn = chrono_rd_index.get(s.lower())
                        if not srn:
                            raise ValueError(f"Chrono name '{nm}' not found in reference index")
                        srns.append(srn)
                robj = {
                    "resqmlType": "resqml20:StratigraphicColumnRankInterpretation",
                    "uuid": rid,
                    "title": r.name,
                    "orderingCriteria": r.ordering,
                    "chronoStratRefs": [{"srn": s} for s in srns],
                }
            rank_objs.append(robj)
            objs.append(robj)

        # Horizons
        for h in self.horizons:
            hobj: Dict[str, Any] = {
                "resqmlType": "resqml20:HorizonInterpretation",
                "uuid": h.uuid,
                "title": h.name,
                "extraMetadata": {
                    "boundaryType": h.boundary_type,
                    "unitName": h.unit_name,
                    "vendor": dict(h.vendor),
                },
            }
            if h.age_ma is not None:
                hobj["ageMa"] = h.age_ma
            objs.append(hobj)

        # Column
        col_obj: Dict[str, Any] = {
            "resqmlType": "resqml20:StratigraphicColumn",
            "uuid": self.vendor.get("uuid") or sanitize(self.name),
            "title": self.name,
            "rankInterpretationRefs": [
                {"uuid": ro["uuid"], "contentType": "resqml20:StratigraphicColumnRankInterpretation"}
                for ro in rank_objs
            ],
        }
        if self.vendor:
            col_obj["extraMetadata"] = {"vendor": dict(self.vendor)}
        objs.append(col_obj)
        return objs

    def to_osdu_bundle(
        self,
        partition: str = DEFAULT_PARTITION,
        chrono_rd_index: Optional[Dict[str, str]] = None,
        vendor_map: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        records: List[Dict[str, Any]] = []
        rank_ids: List[str] = []
        vendor_map = vendor_map or VENDOR_TO_OSDU_MAP_DEFAULT
        unit_id_by_uuid: Dict[str, str] = {}

        # Units for litho ranks
        for r in self.ranks:
            if r.kind != "litho":
                continue
            for u in r.units:
                rec_id = wpc_id(partition, "StratigraphicUnitInterpretation", u.uuid or u.name)
                unit_id_by_uuid[u.uuid] = rec_id
                rec = {
                    "id": rec_id,
                    "kind": KIND_UNIT,
                    "data": {"Name": u.name, "VendorMetadata": {"SourceSystem": "SMDA", "Raw": dict(u.vendor)}},
                }
                # Apply mapping from vendor to OSDU structured fields (opt-in)
                apply_vendor_map(u.vendor, rec, vendor_map)
                # Write model-level fields to ensure they survive round-trips
                if u.top_age_ma is not None:
                    _set_path(rec, "data.TimeRange.TopAgeMa", u.top_age_ma)
                if u.base_age_ma is not None:
                    _set_path(rec, "data.TimeRange.BaseAgeMa", u.base_age_ma)
                if u.parent_name:
                    _set_path(rec, "data.ParentName", u.parent_name)
                if u.color_html:
                    _set_path(rec, "data.Rendering.ColorHtml", u.color_html)
                if u.level is not None:
                    _set_path(rec, "data.VendorMetadata.Raw.Level", u.level)
                records.append(rec)

        # Ranks – scope each rank ID by column name so that different columns
        # never collide on the same generic "Level0 / Level1 / …" identifier.
        for r in self.ranks:
            rank_id_token = f"{self.name}_{r.name}"
            rid = wpc_id(partition, "StratigraphicColumnRankInterpretation", rank_id_token)
            data = {"Name": r.name, "OrderingCriteria": r.ordering}
            if r.kind == "litho":
                data["StratigraphicUnitInterpretationSet"] = [unit_id_by_uuid[u.uuid] for u in r.units]
            else:
                srns: List[str] = []
                for nm in r.chrono_names:
                    s = (nm or "").strip()
                    if ":" in s and s.count(":") >= 2:
                        srns.append(s)
                    else:
                        if not chrono_rd_index:
                            raise ValueError(
                                f"Chrono name '{nm}' requires --chrono-refdata to resolve to SRN"
                            )
                        srn = chrono_rd_index.get(s.lower())
                        if not srn:
                            raise ValueError(f"Chrono name '{nm}' not found in reference index")
                        srns.append(srn)
                data["ChronoStratigraphySet"] = srns
            rec = {"id": rid, "kind": KIND_RANK, "data": data}
            records.append(rec)
            rank_ids.append(rid)

        # Horizons
        for h in self.horizons:
            hid = wpc_id(partition, "HorizonInterpretation", h.uuid or h.name)
            hdata: Dict[str, Any] = {"Name": h.name}
            vm_raw: Dict[str, Any] = dict(h.vendor)
            vm_raw["BoundaryType"] = h.boundary_type
            if h.unit_name:
                vm_raw["UnitName"] = h.unit_name
            if h.age_ma is not None:
                hdata["AgeMa"] = h.age_ma
            if h.unit_id:
                hdata["InterpretedFeatureReferences"] = [h.unit_id]
            elif h.unit_name and unit_id_by_uuid:
                # Try to resolve unit by name
                for uid, uid_id in unit_id_by_uuid.items():
                    if h.unit_name.lower() in uid_id.lower():
                        hdata["InterpretedFeatureReferences"] = [uid_id]
                        break
            hdata["VendorMetadata"] = {"SourceSystem": "StratColumnHandler", "Raw": vm_raw}
            records.append({"id": hid, "kind": KIND_HORIZON, "data": hdata})

        # Column
        cid = wpc_id(partition, "StratigraphicColumn", self.name)
        col_data: Dict[str, Any] = {
            "Name": self.name,
            "StratigraphicColumnRankInterpretationSet": rank_ids,
        }
        if self.vendor:
            col_data["VendorMetadata"] = {"SourceSystem": "SMDA", "Raw": dict(self.vendor)}
        records.append(
            {
                "id": cid,
                "kind": KIND_COL,
                "data": col_data,
            }
        )
        return {"records": records}


# ----------------------------------- CLI -----------------------------------
def _load_map(path: Optional[str]) -> Optional[Dict[str, str]]:
    if not path:
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def cmd_smda2resqml(ns):
    col = StratColumn.from_smda_xlsx(ns.xlsx, sheet=ns.sheet)
    idx = load_reference_index(ns.chrono_refdata)
    objs = col.to_resqml_json(chrono_rd_index=idx)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(objs, f, indent=2)
    print(f"[ok] RESQML JSON written: {ns.output}")


def cmd_smda2osdu(ns):
    col = StratColumn.from_smda_xlsx(ns.xlsx, sheet=ns.sheet)
    idx = load_reference_index(ns.chrono_refdata)
    m = _load_map(ns.vendor_map)
    bundle = col.to_osdu_bundle(partition=ns.partition, chrono_rd_index=idx, vendor_map=m)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2)
    print(f"[ok] OSDU bundle written: {ns.output}")


def cmd_resqml2osdu(ns):
    col = StratColumn.from_resqml_json(ns.resqml_json)
    idx = load_reference_index(getattr(ns, 'chrono_refdata', None))
    m = _load_map(ns.vendor_map)
    bundle = col.to_osdu_bundle(partition=ns.partition, chrono_rd_index=idx, vendor_map=m)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2)
    print(f"[ok] OSDU bundle written: {ns.output}")


def cmd_osdu2resqml(ns):
    col = StratColumn.from_osdu_bundle(ns.manifest)
    objs = col.to_resqml_json(chrono_rd_index=None)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(objs, f, indent=2)
    print(f"[ok] RESQML JSON written: {ns.output}")


def cmd_ow2osdu(ns):
    col = StratColumn.from_openworks_json(ns.json_file)
    idx = load_reference_index(getattr(ns, 'chrono_refdata', None))
    m = _load_map(getattr(ns, 'vendor_map', None))
    bundle = col.to_osdu_bundle(partition=ns.partition, chrono_rd_index=idx, vendor_map=m)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2)
    print(f"[ok] OSDU bundle written from OpenWorks JSON: {ns.output}")


def cmd_smdaapi2osdu(ns):
    col = StratColumn.from_smda_api(
        ns.column,
        base_url=ns.smda_url,
        access_token=ns.token or os.getenv("SMDA_ACCESS_TOKEN"),
        api_key=ns.api_key or os.getenv("SMDA_API_KEY"),
    )
    idx = load_reference_index(getattr(ns, 'chrono_refdata', None))
    m = _load_map(getattr(ns, 'vendor_map', None))
    bundle = col.to_osdu_bundle(partition=ns.partition, chrono_rd_index=idx, vendor_map=m)
    with open(ns.output, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2)
    print(f"[ok] OSDU bundle written from SMDA API: {ns.output}")


def cmd_smdaapi_list(ns):
    names = StratColumn.from_smda_api_list_columns(
        base_url=ns.smda_url,
        access_token=ns.token or os.getenv("SMDA_ACCESS_TOKEN"),
        api_key=ns.api_key or os.getenv("SMDA_API_KEY"),
    )
    for n in names:
        print(n)


def build_parser():
    from textwrap import dedent

    epilog = dedent(r"""
    Examples
    --------
    # 1) SMDA (.xlsx) → RESQML JSON graph
    python stratcolumnhandler.py smda2resqml \
        --xlsx smda-api_strat-units.xlsx \
        --sheet ApiStratUnit \
        -o smda.resqml.json

    # 2) SMDA (.xlsx) → OSDU WPC bundle (partition 'data', default)
    python stratcolumnhandler.py smda2osdu \
        --xlsx smda-api_strat-units.xlsx \
        --sheet ApiStratUnit \
        --partition data \
        -o smda.osdu.json

    # 3) RESQML JSON graph → OSDU WPC bundle
    python stratcolumnhandler.py resqml2osdu \
        --resqml-json smda.resqml.json \
        --partition data \
        -o smda_roundtrip.osdu.json

    # 4) OSDU WPC bundle → RESQML JSON graph
    python stratcolumnhandler.py osdu2resqml \
        --manifest smda.osdu.json \
        -o osdu2.resqml.json

    Notes
    -----
    • Chronostratigraphy:
      - If your columns include chrono ranks (e.g., System/Series), supply one or more
        reference-data files via --chrono-refdata. These files can be:
          { "records": [ { "id": "...", "data": { "Name": "...", "AliasNames": [...], "Code": "..." } }, ... ] }
        or a flat JSON array of records, or a single record. A lookup index is built as:
          lower(Name | AliasNames[] | Code | CodeAsNumber) -> record.id (SRN).
      - When exporting chrono ranks to RESQML or OSDU, names that are not already SRNs are
        resolved using this index. If a name cannot be resolved and no index is provided, the run fails.

    • Vendor-to-OSDU mapping:
      - By default, all vendor/source fields are preserved under:
          data.VendorMetadata = { "SourceSystem": "SMDA", "Raw": { ... } }
      - You may optionally lift selected vendor fields into structured OSDU paths using --vendor-map.
        The map is a JSON object: { "<vendor_key>": "data.<Dotted.Path>", ... }.
        Example:
          {
            "strat_unit_type": "data.VendorMetadata.Raw.UnitType",
            "color_html":     "data.Rendering.ColorHtml",
            "top_age":        "data.TimeRange.TopAgeMa",
            "base_age":       "data.TimeRange.BaseAgeMa"
          }

    • Kinds (WPC):
      - Column:   osdu:wks:work-product-component--StratigraphicColumn:1.2.0
      - Rank:     osdu:wks:work-product-component--StratigraphicColumnRankInterpretation:1.3.0
      - Unit:     osdu:wks:work-product-component--StratigraphicUnitInterpretation:1.3.0
      - Horizon:  osdu:wks:work-product-component--HorizonInterpretation:1.0.0

    • Exit codes: 0 on success; non-zero on Errors.

    • Environment:
      - OSDU_PARTITION may be used to set the default partition (default: 'data').

    • Safety / Round-trip:
      - RESQML extraMetadata reads legacy {"ow": {...}} but writes {"vendor": {...}}.
      - OSDU VendorMetadata reads legacy {"OW": {...}} but writes {"SourceSystem": "SMDA", "Raw": {...}}.
    """)

    ap = argparse.ArgumentParser(
        prog="stratcolumnhandler.py",
        description=dedent("""\
            Strat Column Handler (clean)
            ---------------------------
            Convert stratigraphic columns between:
              • SMDA (.xlsx, sheet 'ApiStratUnit')
              • RESQML 2.0.1 JSON graph
              • OSDU Work-Product-Component (WPC) JSON bundle

            Supports mixed litho/chrono ranks and an arbitrary number of ranks per column.
            """),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=epilog,
    )

    sub = ap.add_subparsers(
        dest="cmd",
        required=True,
        title="Commands",
        metavar="{smda2resqml,smda2osdu,resqml2osdu,osdu2resqml}",
        help="Run 'stratcolumnhandler.py <command> --help' for command-specific options."
    )

    # ---------------------- smda2resqml ----------------------
    p = sub.add_parser(
        "smda2resqml",
        help="SMDA XLSX → RESQML JSON (graph)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Read the SMDA .xlsx (sheet 'ApiStratUnit'), group rows by strat_unit_level,
            and emit a RESQML 2.0.1 JSON graph:
              - resqml20:StratigraphicUnitInterpretation (for litho ranks)
              - resqml20:StratigraphicColumnRankInterpretation (litho and/or chrono)
              - resqml20:StratigraphicColumn

            Chrono ranks: names must resolve to reference-data SRNs; provide --chrono-refdata if needed.
            """),
    )
    p.add_argument(
        "--xlsx", required=True,
        help="Path to the SMDA Excel file (.xlsx) containing the 'ApiStratUnit' sheet."
    )
    p.add_argument(
        "--sheet", default="ApiStratUnit",
        help="Sheet name for SMDA input (default: ApiStratUnit)."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for RESQML JSON graph."
    )
    p.add_argument(
        "--chrono-refdata", nargs="*",
        help="Zero or more JSON file(s) providing ChronoStratigraphy reference-records (SRNs). "
             "Used to resolve chrono rank names to SRNs."
    )
    p.set_defaults(func=cmd_smda2resqml)

    # ---------------------- smda2osdu ------------------------
    p = sub.add_parser(
        "smda2osdu",
        help="SMDA XLSX → OSDU WPC bundle",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Read the SMDA .xlsx and emit an OSDU WPC bundle with:
              - StratigraphicUnitInterpretation records for litho units
              - StratigraphicColumnRankInterpretation records (litho/chrono)
              - StratigraphicColumn record referencing the ranks

            Vendor fields are preserved in data.VendorMetadata.Raw; optionally lift fields using --vendor-map.
            """),
    )
    p.add_argument(
        "--xlsx", required=True,
        help="Path to the SMDA Excel file (.xlsx) containing the 'ApiStratUnit' sheet."
    )
    p.add_argument(
        "--sheet", default="ApiStratUnit",
        help="Sheet name for SMDA input (default: ApiStratUnit)."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for OSDU WPC JSON bundle."
    )
    p.add_argument(
        "--partition", default=DEFAULT_PARTITION,
        help="OSDU partition for record ids (default: env OSDU_PARTITION or 'data')."
    )
    p.add_argument(
        "--chrono-refdata", nargs="*",
        help="Zero or more JSON file(s) providing ChronoStratigraphy reference-records (SRNs)."
    )
    p.add_argument(
        "--vendor-map",
        help="Optional JSON file mapping vendor keys to dotted OSDU paths in the output record. "
             "See help epilog for an example."
    )
    p.set_defaults(func=cmd_smda2osdu)

    # ---------------------- resqml2osdu ----------------------
    p = sub.add_parser(
        "resqml2osdu",
        help="RESQML JSON → OSDU WPC bundle",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Read a RESQML 2.0.1 JSON graph composed of:
              - StratigraphicUnitInterpretation
              - StratigraphicColumnRankInterpretation
              - StratigraphicColumn
            and convert to an OSDU WPC bundle preserving rank order and relationships.

            Vendor payload is taken from extraMetadata.vendor (or legacy extraMetadata.ow) and
            written to data.VendorMetadata.Raw in OSDU; --vendor-map can elevate fields.
            """),
    )
    p.add_argument(
        "--resqml-json", required=True,
        help="Path to the RESQML JSON graph (single list or a single object)."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for OSDU WPC JSON bundle."
    )
    p.add_argument(
        "--partition", default=DEFAULT_PARTITION,
        help="OSDU partition for record ids (default: env OSDU_PARTITION or 'data')."
    )
    p.add_argument(
        "--chrono-refdata", nargs="*",
        help="Zero or more JSON file(s) providing ChronoStratigraphy reference-records (SRNs). "
             "Used to resolve chrono rank names to SRNs."
    )
    p.add_argument(
        "--vendor-map",
        help="Optional JSON file mapping vendor keys to dotted OSDU paths (applied to vendor payload)."
    )
    p.set_defaults(func=cmd_resqml2osdu)

    # ---------------------- osdu2resqml ----------------------
    p = sub.add_parser(
        "osdu2resqml",
        help="OSDU WPC bundle → RESQML JSON (graph)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Read an OSDU WPC JSON bundle containing:
              - StratigraphicUnitInterpretation (units)
              - StratigraphicColumnRankInterpretation (ranks)
              - StratigraphicColumn (column)
            and emit an equivalent RESQML 2.0.1 JSON graph.

            Chrono ranks keep their ChronoStratigraphySet SRNs; no resolution required in this direction.
            """),
    )
    p.add_argument(
        "--manifest", required=True,
        help="Path to the OSDU WPC JSON bundle ('records': [...])."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for RESQML JSON graph."
    )
    p.set_defaults(func=cmd_osdu2resqml)

    # ---------------------- ow2osdu --------------------------
    p = sub.add_parser(
        "ow2osdu",
        help="OpenWorks JSON → OSDU WPC bundle",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Read an OpenWorks JSON export containing:
              {"StratColumn": {"Name": "...", "Type": "litho"}, "Units": [{...}]}
            and emit an OSDU WPC bundle.

            Flexible field name matching: Name/name/identifier, Level/level/strat_unit_level,
            TopAge/topAge/top_age, ParentName/parentName/strat_unit_parent, etc.
            """),
    )
    p.add_argument(
        "--json-file", required=True,
        help="Path to the OpenWorks JSON file."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for OSDU WPC JSON bundle."
    )
    p.add_argument(
        "--partition", default=DEFAULT_PARTITION,
        help="OSDU partition for record ids."
    )
    p.add_argument(
        "--chrono-refdata", nargs="*",
        help="ChronoStratigraphy reference-records for chrono rank resolution."
    )
    p.add_argument(
        "--vendor-map",
        help="Optional JSON mapping vendor keys to OSDU paths."
    )
    p.set_defaults(func=cmd_ow2osdu)

    # ---------------------- smdaapi2osdu ---------------------
    p = sub.add_parser(
        "smdaapi2osdu",
        help="SMDA OPUS API → OSDU WPC bundle (live fetch)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=dedent("""\
            Fetch a stratigraphic column directly from the SMDA API
            (/smda-api/strat-units) and convert to an OSDU WPC bundle.

            Requires authentication: --token (Bearer) or --api-key (Ocp-Apim-Subscription-Key).
            Alternatively set SMDA_ACCESS_TOKEN or SMDA_API_KEY environment variables.
            """),
    )
    p.add_argument(
        "--column", required=True,
        help="strat_column_identifier to fetch (e.g. 'NCS Lithostratigraphy')."
    )
    p.add_argument(
        "--smda-url", default="https://api.gateway.equinor.com",
        help="SMDA API base URL."
    )
    p.add_argument(
        "--token", default="",
        help="Bearer token for Azure AD auth."
    )
    p.add_argument(
        "--api-key", default="",
        help="Ocp-Apim-Subscription-Key."
    )
    p.add_argument(
        "--output", "-o", required=True,
        help="Output path for OSDU WPC JSON bundle."
    )
    p.add_argument(
        "--partition", default=DEFAULT_PARTITION,
        help="OSDU partition for record ids."
    )
    p.add_argument(
        "--chrono-refdata", nargs="*",
        help="ChronoStratigraphy reference-records for chrono rank resolution."
    )
    p.add_argument(
        "--vendor-map",
        help="Optional JSON mapping vendor keys to OSDU paths."
    )
    p.set_defaults(func=cmd_smdaapi2osdu)

    # ---------------------- smdaapi-list ---------------------
    p = sub.add_parser(
        "smdaapi-list",
        help="List available strat column identifiers from SMDA API",
    )
    p.add_argument(
        "--smda-url", default="https://api.gateway.equinor.com",
        help="SMDA API base URL."
    )
    p.add_argument(
        "--token", default="",
        help="Bearer token for Azure AD auth."
    )
    p.add_argument(
        "--api-key", default="",
        help="Ocp-Apim-Subscription-Key."
    )
    p.set_defaults(func=cmd_smdaapi_list)

    return ap


def main():
    ap = build_parser()
    ns = ap.parse_args()
    try:
        ns.func(ns)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()  