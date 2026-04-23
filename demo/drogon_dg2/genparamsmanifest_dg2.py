#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
genparamsmanifest_dg2.py  –  Real Drogon one-by-one sensitivity design matrix
             as an OSDU ColumnBasedTable work-product-component.

Reads the *actual* design_matrix_one_by_one.xlsx produced by fmu-tools
``fmudesign`` for the Drogon FMU v26.0.0 model (drogon_design.ert).

    250 realisation rows  ×  24 columns
    ─────────────────────────────────────
    Key columns : REAL, SENSNAME, SENSCASE
    Value cols  : RMS_SEED + 20 uncertainty parameters

    15 one-by-one sensitivities, each with 10 realisations per case:
      rms_seed (p10_p90), hum (p10_p90), goc (shallow/deep),
      fwl (shallow/deep), fwl_mc (p10_p90), minpv (low),
      kvkh (low/high), faultseal (low/high), relperm (pessim./optim.),
      mulregt (low), multregt_mc (p10_p90),
      valysar_aps_prob_channel (low/high), therys_aps_prob_upshface (low/high),
      therys_aps_prob_loshface (low/high), volon_aps_prob_channel (low/high)

The design-input metadata (distributions, ranges, scenario values) is
embedded in the WPC ``DesignMatrix`` block so consumers can reconstruct
tornado plots without the original xlsx.

Output
------
  manifest_wpcparams_dg2.json

Usage
-----
  python demo/drogon_dg2/genparamsmanifest_dg2.py
  python demo/drogon_dg2/genparamsmanifest_dg2.py --xlsx /path/to/design_matrix_one_by_one.xlsx
"""

import argparse
import json
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import sys

SCRIPT_DIR = Path(__file__).resolve().parent        # demo/drogon_dg2
DG1_DIR    = SCRIPT_DIR.parent / "drogon"            # demo/drogon

if str(DG1_DIR) not in sys.path:
    sys.path.insert(0, str(DG1_DIR))
from _shared import load_json, SEGMENT_NAMES  # noqa: E402


# ── Default path to the cached design matrix ────────────────────────
# The xlsx is copied from the Drogon FMU model repository at:
#   Z:\tutorial\drogon\resmod\ff\26.0.0\ert\input\distributions\
# and cached locally for offline generation.  Override with --xlsx.
DEFAULT_XLSX = "/tmp/design_matrix_one_by_one.xlsx"

# ── Sensitivity definitions from design_input_one_by_one.xlsx ────────
# Each entry: (sensname, type, params_with_ranges)
# type: "seed" | "scenario" | "dist"
# params: list of (param_name, case1, value1, case2, value2, dist, lo, hi)
SENSITIVITY_DEFS: List[Dict[str, Any]] = [
    {
        "SensName": "rms_seed",
        "Type": "seed",
        "NumReal": 10,
        "Parameters": [
            {"Name": "RMS_SEED", "Default": 1000}
        ],
    },
    {
        "SensName": "hum",
        "Type": "dist",
        "NumReal": 20,
        "Parameters": [
            {"Name": "HUM_MODEL_MODE", "Distribution": "const", "Value": 1, "Default": 0}
        ],
    },
    {
        "SensName": "goc",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "GOC_NORTH_HORST", "Case1": "shallow", "Value1": 1635,
             "Case2": "deep", "Value2": 1645, "Default": 1640, "Unit": "m"}
        ],
    },
    {
        "SensName": "fwl",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "FWL_CENTRAL", "Case1": "shallow", "Value1": 1672,
             "Case2": "deep", "Value2": 1682, "Default": 1677, "Unit": "m"},
            {"Name": "FWL_NORTH_HORST", "Case1": "shallow", "Value1": 1655,
             "Case2": "deep", "Value2": 1665, "Default": 1660, "Unit": "m"},
        ],
    },
    {
        "SensName": "fwl_mc",
        "Type": "dist",
        "NumReal": 10,
        "Parameters": [
            {"Name": "FWL_CENTRAL", "Distribution": "unif", "Min": 1672, "Max": 1682, "Default": 1677, "Unit": "m"},
            {"Name": "FWL_NORTH_HORST", "Distribution": "unif", "Min": 1655, "Max": 1665, "Default": 1660, "Unit": "m"},
        ],
    },
    {
        "SensName": "minpv",
        "Type": "scenario",
        "NumReal": 10,
        "Parameters": [
            {"Name": "SIMGRID_MINPV", "Case1": "low", "Value1": 1, "Default": 100, "Unit": "m3"}
        ],
    },
    {
        "SensName": "kvkh",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "KVKH_CHANNEL",  "Case1": "low", "Value1": 0.4, "Case2": "high", "Value2": 0.8, "Default": 0.6},
            {"Name": "KVKH_CREVASSE", "Case1": "low", "Value1": 0.1, "Case2": "high", "Value2": 0.5, "Default": 0.3},
            {"Name": "KVKH_US",       "Case1": "low", "Value1": 0.4, "Case2": "high", "Value2": 0.8, "Default": 0.6},
            {"Name": "KVKH_LS",       "Case1": "low", "Value1": 0.5, "Case2": "high", "Value2": 0.9, "Default": 0.7},
        ],
    },
    {
        "SensName": "faultseal",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "FAULT_SEAL_SCALING", "Case1": "low", "Value1": 0.1, "Case2": "high", "Value2": 10, "Default": 1}
        ],
    },
    {
        "SensName": "relperm",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "RELPERM_INT_WO", "Case1": "pessimistic", "Value1": -1, "Case2": "optimistic", "Value2": 1, "Default": 0},
            {"Name": "RELPERM_INT_GO", "Case1": "pessimistic", "Value1": -1, "Case2": "optimistic", "Value2": 1, "Default": 0},
        ],
    },
    {
        "SensName": "mulregt",
        "Type": "scenario",
        "NumReal": 10,
        "Parameters": [
            {"Name": "MULT_R5_VALYSAR_THERYS", "Case1": "low",  "Value1": 1e-5, "Default": 1},
            {"Name": "MULT_R5_THERYS_VOLON",   "Case1": "low",  "Value1": 1e-5, "Default": 1},
            {"Name": "MULT_R5_THERYS_R2_ALL",  "Case1": "low",  "Value1": 1e-5, "Default": 1},
            {"Name": "MULT_R5_THERYS_R6_ALL",  "Case1": "low",  "Value1": 1e-5, "Default": 1},
        ],
    },
    {
        "SensName": "multregt_mc",
        "Type": "dist",
        "NumReal": 10,
        "Parameters": [
            {"Name": "MULT_R5_VALYSAR_THERYS", "Distribution": "logunif", "Min": 1e-5, "Max": 1, "Default": 1},
            {"Name": "MULT_R5_THERYS_VOLON",   "Distribution": "logunif", "Min": 1e-5, "Max": 1, "Default": 1},
            {"Name": "MULT_R5_THERYS_R2_ALL",  "Distribution": "logunif", "Min": 1e-5, "Max": 1, "Default": 1},
            {"Name": "MULT_R5_THERYS_R6_ALL",  "Distribution": "logunif", "Min": 1e-5, "Max": 1, "Default": 1},
        ],
    },
    {
        "SensName": "valysar_aps_prob_channel",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "VALYSAR_APS_PROB_CHANNEL", "Case1": "low", "Value1": 0.24, "Case2": "high", "Value2": 0.44, "Default": 0.34}
        ],
    },
    {
        "SensName": "therys_aps_prob_upshface",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "THERYS_APS_PROB_UPSHOREFACE", "Case1": "low", "Value1": 0.1, "Case2": "high", "Value2": 0.3, "Default": 0.2}
        ],
    },
    {
        "SensName": "therys_aps_prob_loshface",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "THERYS_APS_PROB_LOWSHOREFACE", "Case1": "low", "Value1": 0.2, "Case2": "high", "Value2": 0.4, "Default": 0.3}
        ],
    },
    {
        "SensName": "volon_aps_prob_channel",
        "Type": "scenario",
        "NumReal": 20,
        "Parameters": [
            {"Name": "VOLON_APS_PROB_CHANNEL", "Case1": "low", "Value1": 0.52, "Case2": "high", "Value2": 0.72, "Default": 0.62}
        ],
    },
]

# ── Column metadata: UoM and value type per design-sheet column ──────
# Columns 3..23 of DesignSheet01 (after REAL, SENSNAME, SENSCASE)
COLUMN_META: Dict[str, Dict[str, str]] = {
    "RMS_SEED":                      {"Unit": "Euc",  "ValueType": "integer"},
    "HUM_MODEL_MODE":                {"Unit": "Euc",  "ValueType": "integer"},
    "GOC_NORTH_HORST":               {"Unit": "m",    "ValueType": "number"},
    "FWL_CENTRAL":                   {"Unit": "m",    "ValueType": "number"},
    "FWL_NORTH_HORST":               {"Unit": "m",    "ValueType": "number"},
    "SIMGRID_MINPV":                 {"Unit": "m3",   "ValueType": "number"},
    "KVKH_CHANNEL":                  {"Unit": "Euc",  "ValueType": "number"},
    "KVKH_CREVASSE":                 {"Unit": "Euc",  "ValueType": "number"},
    "KVKH_LS":                       {"Unit": "Euc",  "ValueType": "number"},
    "KVKH_US":                       {"Unit": "Euc",  "ValueType": "number"},
    "FAULT_SEAL_SCALING":            {"Unit": "Euc",  "ValueType": "number"},
    "RELPERM_INT_GO":                {"Unit": "Euc",  "ValueType": "number"},
    "RELPERM_INT_WO":                {"Unit": "Euc",  "ValueType": "number"},
    "MULT_R5_VALYSAR_THERYS":       {"Unit": "Euc",  "ValueType": "number"},
    "MULT_R5_THERYS_VOLON":         {"Unit": "Euc",  "ValueType": "number"},
    "MULT_R5_THERYS_R2_ALL":        {"Unit": "Euc",  "ValueType": "number"},
    "MULT_R5_THERYS_R6_ALL":        {"Unit": "Euc",  "ValueType": "number"},
    "VALYSAR_APS_PROB_CHANNEL":     {"Unit": "Euc",  "ValueType": "number"},
    "THERYS_APS_PROB_UPSHOREFACE":  {"Unit": "Euc",  "ValueType": "number"},
    "THERYS_APS_PROB_LOWSHOREFACE": {"Unit": "Euc",  "ValueType": "number"},
    "VOLON_APS_PROB_CHANNEL":       {"Unit": "Euc",  "ValueType": "number"},
}


def std_ref_id(prefix: str, entity: str, name: str) -> str:
    return f"{prefix}:reference-data--{entity}:{name}:"

def wpc_id(prefix: str, entity: str, uid: str) -> str:
    return f"{prefix}:work-product-component--{entity}:{uid}:1"


def _read_xlsx(xlsx_path: str) -> tuple:
    """Read DesignSheet01 from the design matrix xlsx.

    Returns (headers, rows) where each row is a list of cell values.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["DesignSheet01"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows.append(row)
    wb.close()
    return headers, rows


def main():
    ap = argparse.ArgumentParser(
        description="Generate DG2 ColumnBasedTable from real Drogon design matrix"
    )
    ap.add_argument("--xlsx",      default=DEFAULT_XLSX,
                    help="Path to design_matrix_one_by_one.xlsx")
    ap.add_argument("--masterwp",  default=str(DG1_DIR / "manifest_masterwp_drogon.json"))
    ap.add_argument("--manifest",  default=str(SCRIPT_DIR / "manifest_wpcparams_dg2.json"))
    ap.add_argument("--id-prefix", default="dev")
    args = ap.parse_args()

    # ── Read the real design matrix ──────────────────────────────────
    headers, rows = _read_xlsx(args.xlsx)
    n_rows = len(rows)
    n_cols = len(headers)
    print(f"Read design matrix: {n_rows} realisations × {n_cols} columns")
    print(f"  Source: {args.xlsx}")

    # Build header→index mapping
    hdr_idx = {h: i for i, h in enumerate(headers)}

    # ── Extract column arrays ────────────────────────────────────────
    realisations: List[int]  = []
    sensnames:    List[str]  = []
    senscases:    List[str]  = []
    param_data:   Dict[str, list] = {h: [] for h in headers[3:]}  # skip REAL/SENSNAME/SENSCASE

    for row in rows:
        realisations.append(int(row[0]))
        sensnames.append(str(row[1]))
        senscases.append(str(row[2]))
        for h in headers[3:]:
            val = row[hdr_idx[h]]
            # Coerce to appropriate type
            if val is None:
                val = 0
            param_data[h].append(val)

    # ── Load master data for ancestry/ACL ────────────────────────────
    masterwp = load_json(args.masterwp)

    reservoir_id = workproduct_id = ""
    acl   = {"owners": [], "viewers": []}
    legal = {"legaltags": [], "otherRelevantDataCountries": []}
    segment_ids: List[str] = []

    for md in masterwp.get("MasterData", []):
        kind = md.get("kind", "")
        if "master-data--Reservoir:" in kind:
            reservoir_id = md["id"]
            acl   = md["acl"]
            legal = md["legal"]
        elif "master-data--ReservoirSegment:" in kind:
            segment_ids.append(md["id"])
    for wp in masterwp.get("Data", {}).get("WorkProducts", []):
        workproduct_id = wp.get("id", "")

    # ── Build KeyColumns ─────────────────────────────────────────────
    key_columns = [
        {"ColumnName": "REAL",      "ColumnRole": "Key", "ValueType": "integer"},
        {"ColumnName": "SENSNAME",  "ColumnRole": "Key", "ValueType": "string"},
        {"ColumnName": "SENSCASE",  "ColumnRole": "Key", "ValueType": "string"},
    ]

    # ── Build Value Columns with metadata ────────────────────────────
    columns = []
    for col_name in headers[3:]:
        meta = COLUMN_META.get(col_name, {"Unit": "Euc", "ValueType": "number"})
        col_def: Dict[str, Any] = {
            "ColumnName":      col_name,
            "ColumnRole":      "Value",
            "ValueType":       meta["ValueType"],
            "UnitOfMeasureID": std_ref_id(args.id_prefix, "UnitOfMeasure", meta["Unit"]),
        }
        columns.append(col_def)

    # ── Build ColumnValues ───────────────────────────────────────────
    column_values: Dict[str, Any] = {
        "REAL":      realisations,
        "SENSNAME":  sensnames,
        "SENSCASE":  senscases,
    }
    for col_name in headers[3:]:
        column_values[col_name] = param_data[col_name]

    # ── Ancestry ─────────────────────────────────────────────────────
    ancestry = {
        "parents":  [reservoir_id] if reservoir_id else [],
        "children": segment_ids,
    }

    # ── Sensitivity summary statistics ───────────────────────────────
    sens_summary: List[Dict[str, Any]] = []
    for sd in SENSITIVITY_DEFS:
        entry: Dict[str, Any] = {
            "SensitivityName": sd["SensName"],
            "SensitivityType": sd["Type"],
            "NumRealisations": sd["NumReal"],
            "Parameters": [],
        }
        for p in sd["Parameters"]:
            pinfo: Dict[str, Any] = {"Name": p["Name"]}
            if "Default" in p:
                pinfo["Default"] = p["Default"]
            if "Distribution" in p:
                pinfo["Distribution"] = p["Distribution"]
            if "Min" in p:
                pinfo["Min"] = p["Min"]
            if "Max" in p:
                pinfo["Max"] = p["Max"]
            if "Value" in p:
                pinfo["Value"] = p["Value"]
            if "Case1" in p:
                pinfo["Case1"] = p["Case1"]
                pinfo["Value1"] = p["Value1"]
            if "Case2" in p:
                pinfo["Case2"] = p["Case2"]
                pinfo["Value2"] = p["Value2"]
            if "Unit" in p:
                pinfo["Unit"] = p["Unit"]
            entry["Parameters"].append(pinfo)
        sens_summary.append(entry)

    # ── WPC record ───────────────────────────────────────────────────
    wpc_record_id = wpc_id(args.id_prefix, "ColumnBasedTable", str(uuid.uuid4()))

    wpc = {
        "id":    wpc_record_id,
        "kind":  "osdu:wks:work-product-component--ColumnBasedTable:1.4.0",
        "acl":   acl,
        "legal": legal,
        "data": {
            "Name": "Drogon DG2 – One-by-one sensitivity design matrix (250 realisations × 21 parameters)",
            "Description": (
                "Real design matrix from the Drogon FMU v26.0.0 one-by-one sensitivity study "
                "(drogon_design.ert). Generated by fmu-tools fmudesign from "
                "design_input_one_by_one.xlsx with 10 seed replicates per case. "
                "15 sensitivities: rms_seed, hum (humidity model), goc (gas-oil contact), "
                "fwl (free-water level scenarios), fwl_mc (FWL Monte Carlo), "
                "minpv (minimum pore-volume), kvkh (vertical/horizontal perm ratio), "
                "faultseal, relperm (relative permeability endpoints), "
                "mulregt/multregt_mc (fault transmissibility multipliers), "
                "and four APS facies probability sensitivities "
                "(Valysar channel, Therys upper/lower shoreface, Volon channel). "
                "250 realisations total. 21 parameter columns with full values. "
                "Design-input metadata (distributions, scenario ranges, defaults) "
                "embedded in the DesignMatrix block."
            ),
            "ParentObjectID":      reservoir_id,
            "ParentWorkProductID": workproduct_id,
            "ancestry": ancestry,
            "DDMSDatasets": [
                f"eml:///dataspace('maap/drogon_dg')/resqml22.TableRepresentation('{wpc_record_id}')"
            ],
            "Table": {
                "ColumnBasedTableTypeID": std_ref_id(
                    args.id_prefix, "ColumnBasedTableType", "AdHoc"
                ),
                "KeyColumns":   key_columns,
                "Columns":      columns,
                "ColumnValues": column_values,
            },
            "DesignMatrix": {
                "Type": "one-by-one sensitivity",
                "Source": "design_matrix_one_by_one.xlsx (drogon_design.ert)",
                "DesignTool": "fmu-tools fmudesign",
                "NumRealisations": n_rows,
                "NumParameters": len(headers) - 3,
                "NumSensitivities": len(SENSITIVITY_DEFS),
                "SeedReplicates": 10,
                "SamplingMethod": "User_Defined",
                "Sensitivities": sens_summary,
            },
        },
    }

    manifest = {
        "kind": "osdu:wks:Manifest:1.0.0",
        "ReferenceData": [],
        "MasterData": [],
        "Data": {
            "Datasets": [],
            "WorkProductComponents": [wpc],
            "WorkProducts": [],
        },
    }

    Path(args.manifest).write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"\nDG2 design-matrix manifest written → {args.manifest}")
    print(f"  WPC ID          : {wpc_record_id}")
    print(f"  Realisations    : {n_rows}")
    print(f"  Parameters      : {len(headers) - 3}")
    print(f"  Sensitivities   : {len(SENSITIVITY_DEFS)}")

    # Print per-sensitivity summary
    from collections import Counter
    sens_counts = Counter(sensnames)
    print(f"\n  Sensitivity breakdown:")
    for sn, cnt in sens_counts.items():
        cases = sorted(set(sc for sn2, sc in zip(sensnames, senscases) if sn2 == sn))
        print(f"    {sn:35s}  {cnt:3d} reals  cases: {', '.join(cases)}")


if __name__ == "__main__":
    main()
